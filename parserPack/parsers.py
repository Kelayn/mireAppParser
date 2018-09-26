import re
import json
from openpyxl import load_workbook
from parserPack import usualDictCreator


def other(dirName, listOfFiles):
    groupNameTmplt = re.compile(r'(?:[а-яё]{4}[-]\d\d[-]\d\d){1}(?:\s[(]\d[)])*(?:[(]\d[)])*', re.I)
    for file in listOfFiles:
        wb = load_workbook(dirName + '\\' +file)
        print(file)
        input()
        sheet = wb.worksheets[0]
        # итерации по колонкам файлам
        for col in sheet.iter_cols(min_row=2, max_row=2):
            cell = col[0]  # Присваивание ячейки
            if cell.value and groupNameTmplt.findall(str(cell.value)):  # Если ячейка не пустая и в ней подходящее значение
                groupName = groupNameTmplt.findall(str(cell.value))
                cell.row += 2
                cell = sheet[cell.column + str(cell.row)]
                evenCell = sheet[cell.column + str(cell.row)]
                schedule = usualDictCreator(sheet, cell, evenCell)
                with open('C:/Test/_json/' + groupName[0] + '.json', 'w') as json_file:
                    json.dump(schedule, json_file, ensure_ascii=False)


def parserChoice(dirName, files):
    if dirName.endswith("\\iep_evening"):
        return  # iep_e(files)
    elif dirName.endswith("\\iep_zaoch"):
        return  # iep_z(files)
    elif dirName.endswith("\\Integu_Zaoch"):
        return  # integu_z(files)
    elif dirName.endswith("\\ITHT_Zaoch"):
        return  # tht_z(files)
    elif dirName.endswith("\\ivzo"):
        return  # ivzo(files)
    elif dirName.endswith("\\main"):
        other(dirName, files)
    elif dirName.endswith("\\main_2"):
        return  # other_2("main_2", files)
    elif dirName.endswith("\\vzo"):
        return  # vzo("vzo", files)
    elif dirName.endswith("\\vzo_2"):
        return  # vzo_2("vzo_2", files)