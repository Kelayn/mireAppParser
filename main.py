from openpyxl import load_workbook
import re
import json
import os
from jsonLoader import jsonLoader
from parserPack import usualWeekHandler



def main():
    # Шаблон на имя группы
    groupNameTmplt = re.compile(r'(?:[а-яё]{4}[-]\d\d[-]\d\d){1}', re.I)
    listOfFiles = os.listdir("C:\Test\schedule")
    for file in listOfFiles:
        wb = load_workbook('C:/Test/schedule/'+file)
        sheet = wb.worksheets[0]
    # итерации по колонкам файлам (по буквам)
        for col in sheet.iter_cols(min_row=2, max_row=2):
            cell = col[0]  # Присваивание ячейки
            if cell.value and groupNameTmplt.findall(str(cell.value)):  # Если ячейка не пустая и в ней подходящее значение
                groupName = groupNameTmplt.findall(str(cell.value))
                print(groupName[0])
                cell.row += 2
                cell.value = sheet[cell.column + str(cell.row)].value
                evenCell = sheet[cell.column + str(cell.row)]
                schedule = {
                    "odd": usualWeekHandler(sheet, cell, 0),
                    "even": usualWeekHandler(sheet, evenCell, 1)
                }
                #  print(schedule)
                with open('C:/Test/json/' + groupName[0] + '.json', 'w') as json_file:
                    json.dump(schedule, json_file, ensure_ascii=False)
    # jsonLoader.uploadJson()


main()